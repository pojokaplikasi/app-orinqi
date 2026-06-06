import openpyxl
import json
from datetime import datetime
import sys

# ================================================================
# Baca data dari Excel file
# ================================================================
def read_excel_jieqi_data(excel_file):
    """Baca data 24 Jieqi dari file Excel."""
    print(f"Membaca data dari {excel_file}...")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Struktur data: {year: {jieqi_name: datetime_string}}
        excel_data = {}
        
        # Asumsi format: kolom A = Year, kolom B-X = 24 Jieqi names
        # Baris pertama adalah header
        headers = []
        for col in range(2, 26):  # Kolom B sampai X (24 Jieqi)
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
            else:
                headers.append(f"Jieqi_{col-1}")
        
        print(f"Ditemukan {len(headers)} Jieqi: {headers[:5]}... (dan {len(headers)-5} lainnya)")
        
        # Baca setiap baris (tahun)
        row_num = 2
        while True:
            year = ws.cell(row=row_num, column=1).value
            if not year:
                break
            
            year_data = {}
            for col_idx, jieqi_name in enumerate(headers, start=2):
                cell_value = ws.cell(row=row_num, column=col_idx).value
                if cell_value:
                    # Konversi ke string format
                    if isinstance(cell_value, datetime):
                        year_data[jieqi_name] = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        year_data[jieqi_name] = str(cell_value)
            
            excel_data[str(year)] = year_data
            row_num += 1
        
        print(f"✅ Berhasil membaca {len(excel_data)} tahun data")
        return excel_data
    
    except Exception as e:
        print(f"❌ Error membaca Excel: {e}")
        return None

# ================================================================
# Baca data dari JSON hasil kalkulasi
# ================================================================
def read_json_calculated_data(json_file):
    """Baca data 23/24 Jieqi dari file JSON hasil kalkulasi."""
    print(f"Membaca data kalkulasi dari {json_file}...")
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"✅ Berhasil membaca {len(data)} tahun data kalkulasi")
        return data
    
    except Exception as e:
        print(f"❌ Error membaca JSON: {e}")
        return None

# ================================================================
# Normalisasi nama Jieqi untuk perbandingan
# ================================================================
def normalize_jieqi_name(name):
    """Normalisasi nama Jieqi agar bisa dibandingkan."""
    name = str(name).lower().strip()
    
    # Mapping variasi nama
    mappings = {
        'lichun': ['lichun', '立春', 'li chun', 'spring begins'],
        'yushui': ['yushui', '雨水', 'yu shui', 'rain water'],
        'jingzhe': ['jingzhe', '惊蛰', 'jing zhe', 'awakening of insects'],
        'chunfen': ['chunfen', '春分', 'chun fen', 'spring equinox'],
        'qingming': ['qingming', '清明', 'qing ming', 'pure brightness'],
        'guyu': ['guyu', '谷雨', 'gu yu', 'grain rain'],
        'lixia': ['lixia', '立夏', 'li xia', 'summer begins'],
        'xiaoman': ['xiaoman', '小满', 'xiao man', 'grain buds'],
        'mangzhong': ['mangzhong', '芒种', 'mang zhong', 'grain in ear'],
        'xiazhi': ['xiazhi', '夏至', 'xia zhi', 'summer solstice'],
        'xiaoshu': ['xiaoshu', '小暑', 'xiao shu', 'slight heat'],
        'dashu': ['dashu', '大暑', 'da shu', 'great heat'],
        'liqiu': ['liqiu', '立秋', 'li qiu', 'autumn begins'],
        'chushu': ['chushu', '处暑', 'chu shu', 'limit of heat'],
        'bailu': ['bailu', '白露', 'bai lu', 'white dew'],
        'qiufen': ['qiufen', '秋分', 'qiu fen', 'autumn equinox'],
        'hanlu': ['hanlu', '寒露', 'han lu', 'cold dew'],
        'shuangjiang': ['shuangjiang', '霜降', 'shuang jiang', 'frost descent'],
        'lidong': ['lidong', '立冬', 'li dong', 'winter begins'],
        'xiaoxue': ['xiaoxue', '小雪', 'xiao xue', 'light snow'],
        'daxue': ['daxue', '大雪', 'da xue', 'heavy snow'],
        'dongzhi': ['dongzhi', '冬至', 'dong zhi', 'winter solstice'],
        'xiaohan': ['xiaohan', '小寒', 'xiao han', 'slight cold'],
        'dahan': ['dahan', '大寒', 'da han', 'great cold'],
    }
    
    # Cari mapping yang cocok
    for standard_name, variations in mappings.items():
        if any(var in name for var in variations):
            return standard_name
    
    # Jika tidak ada mapping, gunakan nama asli
    return name

# ================================================================
# Bandingkan data Excel dengan data kalkulasi
# ================================================================
def compare_datasets(excel_data, calculated_data):
    """Bandingkan data Excel dengan data kalkulasi."""
    print("\n" + "="*80)
    print("MEMULAI PERBANDINGAN DATA")
    print("="*80)
    
    differences = []
    total_compared = 0
    total_same = 0
    total_different = 0
    
    # Ambil tahun yang ada di kedua dataset
    common_years = set(excel_data.keys()) & set(calculated_data.keys())
    print(f"\nTahun yang tersedia di kedua dataset: {len(common_years)} tahun")
    print(f"Range tahun: {min(common_years, key=int)} - {max(common_years, key=int)}")
    
    for year in sorted(common_years, key=int):
        excel_year = excel_data[year]
        calc_year = calculated_data[year]
        
        # Normalisasi kunci untuk perbandingan
        excel_keys_normalized = {normalize_jieqi_name(k): v for k, v in excel_year.items()}
        calc_keys_normalized = {normalize_jieqi_name(k): v for k, v in calc_year.items()}
        
        # Bandingkan setiap Jieqi
        for jieqi_name in excel_keys_normalized:
            if jieqi_name in calc_keys_normalized:
                total_compared += 1
                excel_time = excel_keys_normalized[jieqi_name]
                calc_time = calc_keys_normalized[jieqi_name]
                
                # Normalisasi format waktu untuk perbandingan
                excel_time_norm = excel_time.replace('/', '-').replace('.', '-').strip()
                calc_time_norm = calc_time.replace('/', '-').replace('.', '-').strip()
                
                # Bandingkan
                if excel_time_norm == calc_time_norm:
                    total_same += 1
                else:
                    total_different += 1
                    differences.append({
                        'year': year,
                        'jieqi': jieqi_name,
                        'excel': excel_time,
                        'calculated': calc_time,
                        'difference_seconds': calculate_time_difference(excel_time_norm, calc_time_norm)
                    })
    
    # Tampilkan ringkasan
    print("\n" + "="*80)
    print("RINGKASAN PERBANDINGAN")
    print("="*80)
    print(f"Total Jieqi dibandingkan: {total_compared}")
    print(f"✅ Sama persis: {total_same} ({total_same/total_compared*100:.2f}%)")
    print(f"❌ Berbeda: {total_different} ({total_different/total_compared*100:.2f}%)")
    
    if differences:
        print(f"\n📊 Ditemukan {len(differences)} perbedaan!")
        print("\nDetail perbedaan (maks 20 pertama):")
        print("-"*80)
        for i, diff in enumerate(differences[:20], 1):
            print(f"{i}. Tahun {diff['year']} - {diff['jieqi']}")
            print(f"   Excel:      {diff['excel']}")
            print(f"   Kalkulasi:  {diff['calculated']}")
            print(f"   Selisih:    {diff['difference_seconds']} detik")
            print()
        
        # Simpan semua perbedaan ke file
        save_differences_to_file(differences)
    else:
        print("\n🎉 SEMUA DATA COCOK! Tidak ada perbedaan!")
    
    return differences

# ================================================================
# Hitung selisih waktu dalam detik
# ================================================================
def calculate_time_difference(time1_str, time2_str):
    """Hitung selisih waktu dalam detik antara dua waktu."""
    try:
        # Parse datetime
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y/%m/%d %H:%M:%S',
            '%Y/%m/%d %H:%M',
        ]
        
        dt1 = None
        dt2 = None
        
        for fmt in formats:
            try:
                dt1 = datetime.strptime(time1_str, fmt)
                break
            except:
                continue
        
        for fmt in formats:
            try:
                dt2 = datetime.strptime(time2_str, fmt)
                break
            except:
                continue
        
        if dt1 and dt2:
            diff = abs((dt2 - dt1).total_seconds())
            return int(diff)
        else:
            return None
    
    except Exception as e:
        return None

# ================================================================
# Simpan perbedaan ke file
# ================================================================
def save_differences_to_file(differences):
    """Simpan semua perbedaan ke file untuk analisis lebih lanjut."""
    
    # Simpan sebagai JSON
    with open('comparison_differences.json', 'w', encoding='utf-8') as f:
        json.dump(differences, f, indent=2, ensure_ascii=False)
    print(f"\n💾 Semua perbedaan disimpan ke comparison_differences.json")
    
    # Simpan sebagai TXT yang lebih mudah dibaca
    with open('comparison_differences.txt', 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("PERBEDAAN DATA 24 JIEQI - EXCEL VS KALKULASI\n")
        f.write("="*80 + "\n\n")
        f.write(f"Total perbedaan ditemukan: {len(differences)}\n\n")
        
        # Group by year
        years_with_diff = {}
        for diff in differences:
            year = diff['year']
            if year not in years_with_diff:
                years_with_diff[year] = []
            years_with_diff[year].append(diff)
        
        for year in sorted(years_with_diff.keys(), key=int):
            f.write(f"\n{'='*80}\n")
            f.write(f"TAHUN {year}\n")
            f.write(f"{'='*80}\n")
            
            for diff in years_with_diff[year]:
                f.write(f"\n{diff['jieqi']}:\n")
                f.write(f"  Excel:      {diff['excel']}\n")
                f.write(f"  Kalkulasi:  {diff['calculated']}\n")
                if diff['difference_seconds']:
                    f.write(f"  Selisih:    {diff['difference_seconds']} detik\n")
    
    print(f"💾 Detail perbedaan disimpan ke comparison_differences.txt")

# ================================================================
# Fungsi utama
# ================================================================
def main():
    print("="*80)
    print("PROGRAM PERBANDINGAN DATA 24 JIEQI")
    print("Excel vs Kalkulasi Astronomi (Aplikasi Terintegrasi)")
    print("="*80)
    
    # Check if Excel file exists
    import os
    excel_file = '24_Jieqi_UltraPresisi_VSOP87_Kepler_1909-2183.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"\n⚠️ File Excel '{excel_file}' tidak ditemukan di workspace.")
        print("\n📋 Opsi yang tersedia untuk perbandingan:")
        print("   1. Data dari complete_jieqi_data_1909_2183.json (data Excel yang sudah dikonversi)")
        print("   2. Data dari corrected_final_jieqi_data_1900_2150.json (kalkulasi sebelumnya)")
        print("\n🔍 Mari kita bandingkan kedua dataset ini untuk verifikasi...")
        
        # Compare the two JSON files instead
        excel_data = read_json_calculated_data('complete_jieqi_data_1909_2183.json')
        calculated_data = read_json_calculated_data('corrected_final_jieqi_data_1900_2150.json')
        
        if excel_data and calculated_data:
            print("\n" + "="*80)
            print("PERBANDINGAN: complete_jieqi_data_1909_2183.json vs corrected_final_jieqi_data_1900_2150.json")
            print("="*80)
            differences = compare_datasets(excel_data, calculated_data)
        else:
            print("\n❌ Gagal membaca salah satu file JSON.")
            return
    else:
        # Baca data Excel
        excel_data = read_excel_jieqi_data(excel_file)
        
        if not excel_data:
            print("\n❌ Gagal membaca data Excel. Program dihentikan.")
            return
        
        # Tampilkan sample data Excel
        print("\n📋 Sample data Excel (5 tahun pertama):")
        for i, (year, data) in enumerate(sorted(excel_data.items(), key=lambda x: int(x[0]))):
            if i >= 5:
                break
            print(f"\nTahun {year}:")
            for jieqi, time in list(data.items())[:3]:
                print(f"  {jieqi}: {time}")
            print(f"  ... dan {len(data)-3} Jieqi lainnya")
        
        # Pilih file JSON untuk perbandingan
        print("\n" + "="*80)
        print("PILIH FILE KALKULASI UNTUK PERBANDINGAN")
        print("="*80)
        
        json_files = [f for f in os.listdir('.') if f.endswith('.json') and 'jieqi' in f.lower()]
        
        for i, json_file in enumerate(json_files, 1):
            print(f"  {i}. {json_file}")
        
        # Gunakan complete_jieqi_data_1909_2183.json sebagai default (ini dari Excel)
        selected_json = 'complete_jieqi_data_1909_2183.json'
        print(f"\n🔍 Menggunakan file: {selected_json}")
        
        # Baca data kalkulasi
        calculated_data = read_json_calculated_data(selected_json)
        
        if not calculated_data:
            print("\n❌ Gagal membaca data kalkulasi. Program dihentikan.")
            return
        
        # Lakukan perbandingan
        differences = compare_datasets(excel_data, calculated_data)
    
    print("\n" + "="*80)
    print("✅ PERBANDINGAN SELESAI")
    print("="*80)

if __name__ == "__main__":
    main()
