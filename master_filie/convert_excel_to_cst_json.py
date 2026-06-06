#!/usr/bin/env python3
"""
Convert Excel 24 Jieqi data (WIB UTC+7) to JSON format (CST UTC+8)
and integrate into the Bazi calculator application.
"""

import openpyxl
import json
from datetime import datetime, timedelta

def convert_excel_to_json(excel_file, output_json_file):
    """
    Convert Excel 24 Jieqi data from WIB (UTC+7) to CST (UTC+8)
    and save as JSON file for the Bazi calculator.
    """
    print("="*80)
    print("CONVERTING 24 JIEQI EXCEL DATA (WIB → CST)")
    print("="*80)
    
    # Load Excel file
    print(f"\n📖 Reading Excel file: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    
    # Get the sheet with 24 Jieqi data
    sheet_name = "24 Jieqi WIB (Kompak)"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"✅ Found sheet: {sheet_name}")
    else:
        print(f"❌ Sheet '{sheet_name}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return None
    
    # Skip header rows (first 2 rows are titles)
    # Row 3 contains column headers
    # Data starts from row 4
    
    jieqi_data = {}
    
    # Read data from row 4 onwards
    row_num = 4
    total_years = 0
    
    while True:
        # Column A = Year (Thn)
        year = ws.cell(row=row_num, column=1).value
        
        if not year or str(year).strip() == '':
            break
        
        year = int(year)
        year_data = {}
        
        # Columns: A=Year, B=运(Stem), C=小寒, D=大寒, E=立春, ..., Z=冬至
        # Column mapping based on Excel structure (B is 运, not a Jieqi):
        jieqi_columns = [
            ('xiaohan', 3),    # 小寒 - C
            ('dahan', 4),      # 大寒 - D
            ('lichun', 5),     # 立春 - E
            ('yushui', 6),     # 雨水 - F
            ('jingzhe', 7),    # 惊蛰 - G
            ('chunfen', 8),    # 春分 - H
            ('qingming', 9),   # 清明 - I
            ('guyu', 10),      # 谷雨 - J
            ('lixia', 11),     # 立夏 - K
            ('xiaoman', 12),   # 小满 - L
            ('mangzhong', 13), # 芒种 - M
            ('xiazhi', 14),    # 夏至 - N
            ('xiaoshu', 15),   # 小暑 - O
            ('dashu', 16),     # 大暑 - P
            ('liqiu', 17),     # 立秋 - Q
            ('chushu', 18),    # 处暑 - R
            ('bailu', 19),     # 白露 - S
            ('qiufen', 20),    # 秋分 - T
            ('hanlu', 21),     # 寒露 - U
            ('shuangjiang', 22), # 霜降 - V
            ('lidong', 23),    # 立冬 - W
            ('xiaoxue', 24),   # 小雪 - X
            ('daxue', 25),     # 大雪 - Y
            ('dongzhi', 26),   # 冬至 - Z
        ]
        
        for jieqi_name, col_num in jieqi_columns:
            cell_value = ws.cell(row=row_num, column=col_num).value
            
            if cell_value:
                # Parse the datetime string from Excel
                # Format: "DD MMM\nHH:MM:SS" or similar
                # Need to handle multi-line format
                try:
                    # Convert cell value to string and clean it
                    cell_str = str(cell_value).strip()
                    
                    # Handle different formats
                    # Format 1: "06 Jan\n06:45:12"
                    # Format 2: datetime object
                    if isinstance(cell_value, datetime):
                        dt_wib = cell_value
                    else:
                        # Parse string format
                        # Remove newlines and extra spaces
                        cell_str = cell_str.replace('\n', ' ')
                        
                        # Parse: "06 Jan 06:45:12"
                        # Add year based on row context
                        dt_wib = datetime.strptime(f"{year} {cell_str}", "%Y %d %b %H:%M:%S")
                    
                    # Convert WIB (UTC+7) to CST (UTC+8)
                    # Add 1 hour
                    dt_cst = dt_wib + timedelta(hours=1)
                    
                    # Format as string for JSON
                    year_data[jieqi_name] = dt_cst.strftime('%Y-%m-%d %H:%M:%S')
                    
                except Exception as e:
                    print(f"  ⚠️ Warning: Could not parse {jieqi_name} for year {year}: {cell_value}")
                    print(f"     Error: {e}")
                    # Store as-is if parsing fails
                    year_data[jieqi_name] = str(cell_value)
        
        jieqi_data[str(year)] = year_data
        total_years += 1
        row_num += 1
        
        if total_years % 50 == 0:
            print(f"  Processed {total_years} years...")
    
    print(f"\n✅ Total years processed: {total_years}")
    print(f"   Year range: {min(jieqi_data.keys())} - {max(jieqi_data.keys())}")
    
    # Save to JSON file
    print(f"\n💾 Saving to JSON file: {output_json_file}")
    with open(output_json_file, 'w', encoding='utf-8') as f:
        json.dump(jieqi_data, f, indent=2, ensure_ascii=False)
    
    print(f"✅ JSON file created successfully!")
    
    # Show sample data
    print(f"\n📋 Sample data (first 3 years):")
    for year in sorted(jieqi_data.keys())[:3]:
        print(f"\n  Year {year}:")
        for jieqi, time in list(jieqi_data[year].items())[:4]:
            print(f"    {jieqi}: {time}")
        print(f"    ... and {len(jieqi_data[year])-4} more Jieqi")
    
    return jieqi_data

def main():
    print("\n" + "="*80)
    print("24 JIEQI EXCEL TO JSON CONVERTER")
    print("Timezone: WIB (UTC+7) → CST (UTC+8)")
    print("="*80 + "\n")
    
    excel_file = "24_Jieqi_UltraPresisi_VSOP87_Kepler_1909-2183.xlsx"
    output_json = "jieqi_ultra_precise_1909_2183_cst.json"
    
    # Convert Excel to JSON
    jieqi_data = convert_excel_to_json(excel_file, output_json)
    
    if jieqi_data:
        print("\n" + "="*80)
        print("CONVERSION COMPLETE!")
        print("="*80)
        print(f"\n📁 Output file: {output_json}")
        print(f"📊 Total years: {len(jieqi_data)}")
        print(f"📅 Year range: {min(jieqi_data.keys())} - {max(jieqi_data.keys())}")
        print(f"🎯 Jieqi per year: {len(jieqi_data[list(jieqi_data.keys())[0]])}")
        
        print("\n" + "="*80)
        print("NEXT STEPS:")
        print("="*80)
        print("1. Update app.py to use this new JSON file")
        print("2. Replace COMPLETE_JIEQI_DATA loading with this file")
        print("3. Test the application to verify accuracy")
        print("="*80)

if __name__ == "__main__":
    main()
